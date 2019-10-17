#! /usr/bin/env python3
# -*- coding:utf8 -*-

# PIC-REPORT Ver0.0.1Alpha created 10/10/2019 by S.Wang@Beijing
# See README.md for more information
# BATTLE CONTROL ONLINE

import os  # show help information
import sys  # load system argument package
import time  # load for sleep()

import openpyxl  # load xlsx file r/w package

filename = sys.path[0] + '/record.xlsx'
# sys.path allows linked app to get real path of the origin file
# uncomment to test print(filename)
workbook = openpyxl.load_workbook(filename)
# show sheet name in the workbook
worksheet = workbook[workbook.sheetnames[0]]
# select the active worksheet


# function tested ok 10/10/2019
def record_add():
    patient_id = input("Patient ID:")
    for m in range(1,worksheet.max_row+1):
        if str(worksheet['A'+str(m)].value) == patient_id:
            print('Duplicate ID detected, ID modified automatically')
            patient_id = patient_id + '*'
    patient_name = input('Patient Name:')
    report_status = 'N'

    worksheet.append([patient_id,patient_name,report_status])
    # append data to active sheet
    workbook.save(filename)
    # VERY IMPORTANT, otherwise the data will be lost

    print('\nAdd new record:\n')
    print('ID'.ljust(20),end='')
    print('Name'.ljust(20),end='')
    print('Status'.ljust(20))
    print(patient_id.ljust(20),end='')
    print(patient_name.ljust(20),end='')
    print(report_status.ljust(20)+'\n')
    print('Success!\n')

# function tested ok 11/10/2019
def record_show():

        print()
        print('ID'.ljust(20),end='')
        print('Name'.ljust(20),end='')
        print('Status'.ljust(20))
        
        stat = 0

        for m in range(1,worksheet.max_row+1): # search for each row
            if worksheet['C'+str(m)].value == 'N': # show only record without N mark
                for n in worksheet[m]: # search for each column
                    print(str(n.value).ljust(20),end='')
                    stat =1
                print() # make a seperation for each record
        if stat == 0:
            print('no record remains unsubmitted!\n')
        print()

# function tested ok 10/10/2019
def record_showall():

    print()
    print('ID'.ljust(20),end='')
    print('Name'.ljust(20),end='')
    print('Status'.ljust(20))
    
    for m in range(1,worksheet.max_row+1): # search for each row
        if worksheet['C'+str(m)].value != 'DEL': # show only record without DEL mark
            for n in worksheet[m]: # search for each column
                print(str(n.value).ljust(20),end='')
            print() # make a seperation for each record
    print()

# function tested ok 10/10/2019
def record_showfull():

    print()
    print('ID'.ljust(20),end='')
    print('Name'.ljust(20),end='')
    print('Status'.ljust(20))
    
    for m in range(1,worksheet.max_row+1): # search for each row
        for n in worksheet[m]: # search for each column
            print(str(n.value).ljust(20),end='')
        print() # make a seperation for each record
    print()

# function tested ok 10/10/2019
def record_switch(id):

    runstat = 0
    for m in range(1,worksheet.max_row+1):
        if str(worksheet['A'+str(m)].value) == id and str(worksheet['C'+str(m)].value) != 'DEL':
            print('\nOld Record')
            print('ID'.ljust(20),end='')
            print('Name'.ljust(20),end='')
            print('Status'.ljust(20))
            print(str(worksheet['A'+str(m)].value).ljust(20),end='')
            print(str(worksheet['B'+str(m)].value).ljust(20),end='')
            print(str(worksheet['C'+str(m)].value).ljust(1),end='')
            if worksheet['C'+str(m)].value == 'Y':
                worksheet['C'+str(m)].value = 'N'
            elif worksheet['C'+str(m)].value == 'N':
                worksheet['C'+str(m)].value = 'Y'

            workbook.save(filename)
            
            print('===>'.ljust(0),end='')
            print(str(worksheet['C'+str(m)].value).ljust(1))
            print()
            runstat = 1 # mark the status of switch
        elif str(worksheet['A'+str(m)].value) == id and str(worksheet['C'+str(m)].value) == 'DEL':
            print('DELETED record could not be switched, use undelete to retrive')
            runstat = 1

    if runstat == 0:
        print('NO RECORD SWITCHED, PLEASE DOUBLE CHECK THE ID!')

# function tested ok 10/10/2019
def record_del(id):
    runstat = 0
    print('ID'.ljust(20),end='')
    print('Name'.ljust(20),end='')
    print('Status'.ljust(20))

    for m in range(1,worksheet.max_row+1):
        if str(worksheet['A'+str(m)].value) == id:
            print(str(worksheet['A'+str(m)].value).ljust(20),end='')
            print(str(worksheet['B'+str(m)].value).ljust(20),end='')
            print(str(worksheet['C'+str(m)].value).ljust(5),end='')
            if worksheet['C'+str(m)].value != 'DEL':
                worksheet['C'+str(m)].value = 'DEL'
                workbook.save(filename)
                print('===>',end='')
                print('Deleted')
                runstat = 1 # mark the status of switch
            else:
                print('\nThe record has already been deleted!\n')
                runstat = 1

    if runstat == 0:
        print('NO RECORD FIND, PLEASE DOUBLE CHECK THE ID!')
   
# function tested ok 10/10/2019
def record_undel(id):
    runstat = 0
    print('ID'.ljust(20),end='')
    print('Name'.ljust(20),end='')
    print('Status'.ljust(20))
    for m in range(1,worksheet.max_row+1):
        if str(worksheet['A'+str(m)].value) == id:
            print(str(worksheet['A'+str(m)].value).ljust(20),end='')
            print(str(worksheet['B'+str(m)].value).ljust(20),end='')
            print(str(worksheet['C'+str(m)].value).ljust(5),end='')
            if worksheet['C'+str(m)].value == 'DEL':
                worksheet['C'+str(m)].value = 'N'
                workbook.save(filename)
                print('===>',end='')
                print('Undeleted')
                runstat = 1 # mark the status of switch
            else:
                print('\nThe record has not been deleted!\n')
                runstat = 1


    if runstat == 0:
        print('NO RECORD FIND, PLEASE DOUBLE CHECK THE ID!')

# function tested ok 10/10/2019
def record_applydel():
    stat = 0
    confirmation = input('You will DELETE all records marked as DEL forever, confirm(y/n)?')
    if confirmation == 'y':
        time.sleep(0.5)
        worksheet=workbook[workbook.sheetnames[0]]
        print('ID'.ljust(20),end='')
        print('Name'.ljust(20))
        for m in range(1,worksheet.max_row+1):
            #worksheet = workbook[workbook.sheetnames[0]] # for sometime, you need to reload the worksheet for unknown reason, otherwise the max_row will go wrong
            for n in range (1,worksheet.max_row+1):
                if str(worksheet['C'+str(m)].value) == 'DEL':
                    print(str(worksheet['A'+str(m)].value).ljust(20),end='')
                    print(str(worksheet['B'+str(m)].value).ljust(20) + '===>',end='    ')
                    print('DELETED!!!')
                    worksheet.delete_rows(m)
                    workbook.save(filename)
                    time.sleep(0.5)
                    stat = 1
    if stat == 0:
        print('NOTHING DELETED')


try:
    parameter = sys.argv[1]
    if parameter == 'add':
        record_add()
    elif parameter == 'show':
        record_show()
    elif parameter == 'showall':
        record_showall()
    elif parameter == 'showfull':
        record_showfull()
    elif parameter == 'change':
        try:
            record_switch(sys.argv[2])
        except IndexError:
            print('please input ID you want to switch')
    elif parameter == 'del':
        try:
            record_del(sys.argv[2])
        except IndexError:
            print('please input ID you want to switch')
    elif parameter == 'undel':
        try:
            record_undel(sys.argv[2])
        except IndexError:
            print('please input ID you want to switch')
    elif parameter == 'erase':
        record_applydel()
    elif parameter == '-help':
        os.system('cat '+str(sys.path[0]+'/README.md | less'))
    else:
        print('wrong argument, see README.MD for help')
except KeyboardInterrupt:
    print('\nForced exit by user')
except IndexError:
    input('Press any key to show help information. You can also see help by typing pcir -help')
    os.system('cat '+str(sys.path[0]+'/README.md | less'))
